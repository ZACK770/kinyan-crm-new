"""
סקריפט לתיקון תאריכי יצירה של לידים שיובאו מאקסל
מתקן את created_at ו-arrival_date לפי התאריך האמיתי מהאקסל המקורי
"""
import asyncio
import asyncpg
from datetime import datetime
import openpyxl

# פרטי חיבור ל-PostgreSQL (Render Production)
DB_URL = "postgresql://crm_new_user:45RsFRWnUuvPQFAttG37PxisVlC79HZv@dpg-d65jjr56ubrc7396u8r0-a.frankfurt-postgres.render.com/crm_new"

# נתיב לקובץ אקסל
EXCEL_PATH = r"C:\Users\admin\Downloads\לידים_02_11_2026.xlsx"


def parse_date_from_excel(date_str):
    """פרסור תאריך מפורמט DD/MM/YYYY HH:MM:SS"""
    if not date_str:
        return None
    
    try:
        # פורמט מהאקסל: "11/02/2026 12:25:52"
        return datetime.strptime(str(date_str).strip(), "%d/%m/%Y %H:%M:%S")
    except ValueError:
        try:
            # ניסיון בלי שניות
            return datetime.strptime(str(date_str).strip(), "%d/%m/%Y %H:%M")
        except ValueError:
            try:
                # ניסיון רק תאריך
                return datetime.strptime(str(date_str).strip(), "%d/%m/%Y")
            except ValueError:
                print(f"⚠️  לא הצלחתי לפרסר תאריך: {date_str}")
                return None


async def fix_dates(dry_run=True):
    """
    תיקון תאריכי יצירה של לידים מהאקסל
    
    Args:
        dry_run: אם True, רק מציג מה יעשה בלי לעדכן בפועל
    """
    print("=" * 80)
    print("תיקון תאריכי יצירה של לידים מייבוא אקסל")
    print("=" * 80)
    
    if dry_run:
        print("\n⚠️  מצב DRY RUN - לא יעדכן כלום ב-DB")
    else:
        print("\n🔴 מצב LIVE - יעדכן את ה-DB!")
    
    # 1. קריאת האקסל
    print(f"\n📂 קורא אקסל: {EXCEL_PATH}")
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        
        print(f"✅ נמצאו {ws.max_row - 1} שורות באקסל")
        
        # מציאת עמודות
        phone_col_idx = headers.index("טלפון ראשי") + 1
        date_col_idx = headers.index("תאריך יצירה") + 1
        name_col_idx = headers.index("שם מלא") + 1 if "שם מלא" in headers else None
        
    except FileNotFoundError:
        print(f"❌ קובץ לא נמצא: {EXCEL_PATH}")
        return
    except Exception as e:
        print(f"❌ שגיאה בקריאת אקסל: {e}")
        return
    
    # 2. בניית מיפוי טלפון -> תאריך
    print(f"\n📊 בונה מיפוי טלפון -> תאריך...")
    phone_to_date = {}
    phone_to_name = {}
    dates_found = 0
    dates_missing = 0
    
    for row_idx in range(2, ws.max_row + 1):
        phone = ws.cell(row_idx, phone_col_idx).value
        date_str = ws.cell(row_idx, date_col_idx).value
        name = ws.cell(row_idx, name_col_idx).value if name_col_idx else None
        
        if phone:
            phone = str(phone).strip()
            phone_to_name[phone] = name or "ללא שם"
            
            if date_str:
                parsed_date = parse_date_from_excel(date_str)
                if parsed_date:
                    phone_to_date[phone] = parsed_date
                    dates_found += 1
                else:
                    dates_missing += 1
            else:
                dates_missing += 1
    
    print(f"✅ נמצאו {dates_found} תאריכים תקינים")
    print(f"⚠️  {dates_missing} שורות ללא תאריך תקין")
    
    # 3. חיבור ל-DB
    print(f"\n🔌 מתחבר ל-PostgreSQL...")
    try:
        conn = await asyncpg.connect(DB_URL)
        print("✅ התחברות הצליחה")
    except Exception as e:
        print(f"❌ שגיאת חיבור: {e}")
        return
    
    try:
        # 4. מציאת לידים שצריך לעדכן
        print(f"\n🔍 מחפש לידים עם created_at ב-11/2/2026...")
        
        query = """
            SELECT id, full_name, phone, created_at, arrival_date
            FROM leads
            WHERE created_at >= '2026-02-11 00:00:00'
              AND created_at < '2026-02-12 00:00:00'
            ORDER BY id
        """
        
        leads = await conn.fetch(query)
        print(f"✅ נמצאו {len(leads)} לידים עם created_at ב-11/2/2026")
        
        # 5. עדכון לידים
        stats = {
            "updated": 0,
            "not_in_excel": 0,
            "no_date_in_excel": 0,
            "already_correct": 0,
        }
        
        print(f"\n📝 מעבד לידים...")
        print("-" * 80)
        
        for lead in leads:
            lead_id = lead['id']
            phone = lead['phone']
            name = lead['full_name']
            current_created = lead['created_at']
            current_arrival = lead['arrival_date']
            
            # בדיקה אם הטלפון קיים באקסל
            if phone not in phone_to_date:
                stats["not_in_excel"] += 1
                print(f"⏭️  ID {lead_id}: {name} ({phone}) - לא נמצא באקסל")
                continue
            
            # קבלת התאריך הנכון מהאקסל
            correct_date = phone_to_date[phone]
            
            # בדיקה אם כבר נכון
            if (current_created.replace(tzinfo=None) == correct_date and 
                current_arrival.replace(tzinfo=None) == correct_date):
                stats["already_correct"] += 1
                print(f"✅ ID {lead_id}: {name} ({phone}) - כבר נכון")
                continue
            
            # עדכון
            print(f"\n🔧 ID {lead_id}: {name} ({phone})")
            print(f"   created_at:   {current_created} → {correct_date}")
            print(f"   arrival_date: {current_arrival} → {correct_date}")
            
            if not dry_run:
                update_query = """
                    UPDATE leads
                    SET created_at = $1, arrival_date = $2
                    WHERE id = $3
                """
                await conn.execute(update_query, correct_date, correct_date, lead_id)
                stats["updated"] += 1
            else:
                stats["updated"] += 1
        
        # 6. סיכום
        print("\n" + "=" * 80)
        print("סיכום:")
        print("=" * 80)
        print(f"✅ עודכנו: {stats['updated']}")
        print(f"✓  כבר נכונים: {stats['already_correct']}")
        print(f"⏭️  לא נמצאו באקסל: {stats['not_in_excel']}")
        print(f"📊 סה\"כ לידים: {len(leads)}")
        
        if dry_run:
            print("\n⚠️  זה היה DRY RUN - לא עודכן כלום")
            print("להרצה אמיתית: הרץ עם --live")
        else:
            print("\n✅ עדכון הושלם בהצלחה!")
        
    finally:
        await conn.close()
        print("\n🔌 חיבור נסגר")


async def main():
    import argparse
    parser = argparse.ArgumentParser(description="תיקון תאריכי יצירה מייבוא אקסל")
    parser.add_argument("--live", action="store_true", help="הרצה אמיתית (לא dry-run)")
    args = parser.parse_args()
    
    await fix_dates(dry_run=not args.live)


if __name__ == "__main__":
    asyncio.run(main())
