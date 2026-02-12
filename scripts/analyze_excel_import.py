"""Analyze Excel file for lead import — check headers, data types, values"""
import openpyxl
import sys
from collections import Counter

path = r"C:\Users\משתמש\Documents\Downloads\_הורדה-לסנכרון-12-2-2026-22.30_02_12_2026.xlsx"
wb = openpyxl.load_workbook(path)
ws = wb.active

headers = [cell.value for cell in ws[1]]
total_rows = ws.max_row - 1

print("=" * 70)
print(f"HEADERS ({len(headers)} columns):")
print("=" * 70)
for i, h in enumerate(headers):
    print(f"  [{i}] {h}")

print(f"\nTotal data rows: {total_rows}")

# Sample first 5 rows
print("\n" + "=" * 70)
print("SAMPLE ROWS (first 5):")
print("=" * 70)
for row_idx in range(2, min(7, ws.max_row + 1)):
    print(f"\n--- Row {row_idx - 1} ---")
    for i, h in enumerate(headers):
        val = ws.cell(row_idx, i + 1).value
        if val is not None:
            print(f"  {h}: {repr(val)} (type: {type(val).__name__})")

# Analyze unique values for key columns
print("\n" + "=" * 70)
print("UNIQUE VALUES ANALYSIS:")
print("=" * 70)

key_columns = ["סטאטוס ליד", "סטטוס מענה", "איש מכירות ", "מוצר שמתעניין", "שם המפרסם"]
# Also try without trailing space
alt_columns = ["איש מכירות"]

for col_name in key_columns + alt_columns:
    if col_name in headers:
        col_idx = headers.index(col_name) + 1
        values = Counter()
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row_idx, col_idx).value
            if val is not None:
                values[str(val).strip()] += 1
            else:
                values["(ריק)"] += 1
        print(f"\n  [{col_name}] ({sum(values.values())} values):")
        for val, count in values.most_common():
            print(f"    {val}: {count}")

# Check date columns
print("\n" + "=" * 70)
print("DATE COLUMNS ANALYSIS:")
print("=" * 70)
date_columns = ["תאריך יצירה", "תאריך פניה אחרונה"]
for col_name in date_columns:
    if col_name in headers:
        col_idx = headers.index(col_name) + 1
        types = Counter()
        sample_values = []
        null_count = 0
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row_idx, col_idx).value
            if val is None:
                null_count += 1
            else:
                types[type(val).__name__] += 1
                if len(sample_values) < 5:
                    sample_values.append(repr(val))
        print(f"\n  [{col_name}]:")
        print(f"    Null: {null_count}")
        for t, count in types.most_common():
            print(f"    Type {t}: {count}")
        print(f"    Samples: {sample_values}")

# Check phone format
print("\n" + "=" * 70)
print("PHONE ANALYSIS:")
print("=" * 70)
phone_col = "טלפון ראשי"
if phone_col in headers:
    col_idx = headers.index(phone_col) + 1
    types = Counter()
    null_count = 0
    samples = []
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row_idx, col_idx).value
        if val is None:
            null_count += 1
        else:
            types[type(val).__name__] += 1
            if len(samples) < 5:
                samples.append(repr(val))
    print(f"  Null: {null_count}")
    for t, count in types.most_common():
        print(f"  Type {t}: {count}")
    print(f"  Samples: {samples}")
