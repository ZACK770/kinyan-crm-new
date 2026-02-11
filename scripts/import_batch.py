import asyncio
import sys
sys.path.insert(0, r'c:\Users\admin\kinyan-crm-new')
import openpyxl
from datetime import datetime
from db import SessionLocal
from db.models import Lead, Salesperson, Course
from sqlalchemy import select

SALESPERSON_MAPPING = {"שרוליק": "ישראל ברים", "שלוימי גרוס": "שלמה גרוס", "אהרן מאירוביץ": "אהרן מאירוביץ", "משה גרינהויז": "משה גרינהויז", "נתנאל גפנר": "נתנאל גפנר", "שלמה דנציגר": "שלמה דנציגר", "N/A": None}
COURSE_MAPPING = {"הלכות שבת": "שבת", "ממונות (חושן משפט)": None, "הלכות נידה/טהרה": "טהרה", "איסור והיתר": "איסור והיתר", "מסלול קניין שבת": "שבת", "השלים מבחן - טרם בחר הטבה": None, "מתעניין במסלול::": None}
STATUS_MAPPING = {"ליד חדש": "ליד חדש", "ליד בתהליך": "ליד בתהליך", "חיוג ראשון": "ליד חדש", "לא רלוונטי": "לא רלוונטי", "ליד סגור - לקוח": "ליד סגור - לקוח"}
RESPONSE_MAPPING = {"נענה": "מעוניין", "ניתוק": "לא זמין", "לא נענה (Timeout)": "לא זמין", "פניה כללית": None}

def parse_date(d):
    if not d: return None
    for fmt in ["%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y"]:
        try: return datetime.strptime(str(d), fmt)
        except: pass
    return None

async def import_batch(start, end):
    wb = openpyxl.load_workbook(r"C:\Users\admin\Downloads\לידים_02_11_2026.xlsx")
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    
    async with SessionLocal() as session:
        sp_res = await session.execute(select(Salesperson))
        sp_map = {sp.name: sp.id for sp in sp_res.scalars()}
        c_res = await session.execute(select(Course))
        c_map = {c.name: c.id for c in c_res.scalars()}
        
        stats = {"ok": 0, "dup": 0, "skip": 0}
        
        for row_idx in range(start + 1, min(end + 2, ws.max_row + 1)):
            row_data = {headers[i]: ws.cell(row_idx, i + 1).value for i in range(len(headers)) if headers[i]}
            phone = row_data.get("טלפון ראשי")
            if not phone:
                stats["skip"] += 1
                continue
            phone = str(phone).strip()
            
            ex = await session.execute(select(Lead).where(Lead.phone == phone))
            if ex.scalar_one_or_none():
                stats["dup"] += 1
                continue
            
            try:
                sp_name = row_data.get("איש מכירות ")
                sp_id = sp_map.get(SALESPERSON_MAPPING.get(sp_name.strip())) if sp_name else None
                c_name = row_data.get("מוצר שמתעניין")
                c_id = c_map.get(COURSE_MAPPING.get(c_name.strip())) if c_name else None
                resp = row_data.get("סטטוס מענה")
                
                lead = Lead(
                    full_name=row_data.get("שם מלא") or "ליד ללא שם",
                    phone=phone,
                    email=row_data.get("מייל לקוח"),
                    city=row_data.get("עיר מגורים"),
                    address=row_data.get("כתובת"),
                    notes=row_data.get("הערות ליד"),
                    source_type="ייבוא ממערכת ישנה",
                    source_message=row_data.get("הודעה מהליד"),
                    source_name=row_data.get("שלוחת מוצר"),
                    campaign_name=row_data.get("שם המפרסם"),
                    arrival_date=parse_date(row_data.get("תאריך יצירה")) or datetime.now(),
                    last_contact_date=parse_date(row_data.get("תאריך פניה אחרונה")),
                    status=STATUS_MAPPING.get(row_data.get("סטאטוס ליד", "ליד חדש"), "ליד חדש"),
                    lead_response=RESPONSE_MAPPING.get(resp.strip()) if resp else None,
                    salesperson_id=sp_id,
                    course_id=c_id,
                    created_by="import_script"
                )
                session.add(lead)
                stats["ok"] += 1
            except:
                stats["skip"] += 1
        
        await session.commit()
        print(f"Batch {start}-{end}: ✅{stats['ok']} ⏭️{stats['dup']} ❌{stats['skip']}")

async def main():
    batches = [(1,50), (51,100), (101,150), (151,200), (201,250), (251,300), (301,350), (351,400), (401,469)]
    for start, end in batches:
        await import_batch(start, end)
    print("✅ סיים ייבוא כל ה-batches")

if __name__ == "__main__":
    asyncio.run(main())
