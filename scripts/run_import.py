"""Run lead import directly (no server needed) — merge mode with created_at update"""
import asyncio, sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from api.import_api import normalize_phone, _get_cell, parse_date, SALESPERSON_MAPPING, STATUS_MAPPING, RESPONSE_MAPPING, COURSE_MAPPING
from datetime import datetime, timezone

EXCEL_PATH = r"C:\Users\משתמש\Documents\Downloads\_הורדה-לסנכרון-12-2-2026-22.30_02_12_2026.xlsx"

async def main():
    import openpyxl
    from db import SessionLocal
    from db.models import Lead, Salesperson, Course
    from sqlalchemy import select

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    total_rows = ws.max_row - 1

    print(f"Total rows in Excel: {total_rows}")

    stats = {"created": 0, "merged": 0, "skipped_no_phone": 0, "errors": 0}
    error_details = []

    async with SessionLocal() as session:
        # Load mappings
        sp_res = await session.execute(select(Salesperson))
        sp_ids = {sp.name: sp.id for sp in sp_res.scalars()}
        c_res = await session.execute(select(Course))
        c_ids = {c.name: c.id for c in c_res.scalars()}

        print(f"Salespeople in DB: {len(sp_ids)}")
        print(f"Courses in DB: {len(c_ids)}")
        print()

        for row_idx in range(2, ws.max_row + 1):
            row = {}
            for i, h in enumerate(headers):
                if h:
                    row[h] = ws.cell(row_idx, i + 1).value

            phone_raw = row.get("טלפון ראשי")
            phone = normalize_phone(phone_raw)
            if not phone:
                stats["skipped_no_phone"] += 1
                continue

            try:
                full_name = row.get("שם מלא") or "ליד ללא שם"
                family_name = row.get("משפחה")
                sp_n = _get_cell(row, "איש מכירות", "איש מכירות ")
                sp_mapped = SALESPERSON_MAPPING.get(sp_n.strip()) if sp_n else None
                sp_id = sp_ids.get(sp_mapped) if sp_mapped else None
                requested_course = _get_cell(row, "מוצר שמתעניין", "מוצר", "שם קורס", "קורס")
                if requested_course:
                    requested_course = str(requested_course).strip()
                c_id = None
                if requested_course:
                    mapped_course = COURSE_MAPPING.get(requested_course)
                    if mapped_course:
                        c_id = c_ids.get(mapped_course)
                resp = _get_cell(row, "סטטוס מענה")
                lead_response = RESPONSE_MAPPING.get(resp.strip()) if resp else None
                status_raw = _get_cell(row, "סטאטוס ליד") or "ליד חדש"
                status = STATUS_MAPPING.get(str(status_raw).strip(), "ליד חדש")
                created_date = parse_date(_get_cell(row, "תאריך יצירה", "תאריך הגעה")) or datetime.now(timezone.utc)
                last_contact = parse_date(_get_cell(row, "תאריך פניה אחרונה"))
                phone2 = normalize_phone(row.get("טלפון נוסף"))
                id_number = row.get("תעודת זהות")
                if id_number:
                    id_number = str(id_number).strip()
                notes_parts = []
                for nk in ["הערות ליד", "הערות על הליד"]:
                    v = row.get(nk)
                    if v and str(v).strip():
                        notes_parts.append(str(v).strip())
                notes = "\n".join(notes_parts) if notes_parts else None

                new_data = dict(
                    full_name=full_name,
                    family_name=family_name,
                    phone=phone,
                    phone2=phone2,
                    email=row.get("מייל לקוח"),
                    city=row.get("עיר מגורים"),
                    address=row.get("כתובת"),
                    id_number=id_number,
                    notes=notes,
                    source_type=_get_cell(row, "מקור הגעה כללי") or "ייבוא ממערכת ישנה",
                    source_message=row.get("הודעה מהליד"),
                    campaign_name=_get_cell(row, "שם המפרסם", "שם קמפיין", "קמפיין"),
                    requested_course=requested_course,
                    arrival_date=created_date,
                    last_contact_date=last_contact,
                    status=status,
                    lead_response=lead_response,
                    salesperson_id=sp_id,
                    course_id=c_id,
                    created_at=created_date,
                    created_by="import_script",
                )
            except Exception as e:
                stats["errors"] += 1
                error_details.append({"row": row_idx - 1, "error": str(e)})
                continue

            # Check for existing lead by phone
            ex_result = await session.execute(select(Lead).where(Lead.phone == phone))
            existing = ex_result.scalar_one_or_none()

            if existing:
                # MERGE — update empty fields + always update created_at
                for key, val in new_data.items():
                    if val is not None and key != "phone":
                        current = getattr(existing, key, None)
                        if current is None or current == "" or current == "ליד ללא שם":
                            setattr(existing, key, val)
                # created_at — always update from Excel (real creation date)
                if new_data.get("created_at"):
                    existing.created_at = new_data["created_at"]
                    existing.arrival_date = new_data["created_at"]
                # Notes — append
                if new_data.get("notes") and existing.notes:
                    if new_data["notes"] not in existing.notes:
                        existing.notes = existing.notes + "\n---\n" + new_data["notes"]
                stats["merged"] += 1
            else:
                # New lead
                session.add(Lead(**new_data))
                stats["created"] += 1

            # Progress every 500
            processed = stats["created"] + stats["merged"] + stats["skipped_no_phone"] + stats["errors"]
            if processed % 500 == 0:
                print(f"  Progress: {processed}/{total_rows} ...")

        await session.commit()
        print("\n💾 Committed to DB!")

    print("\n" + "=" * 60)
    print("IMPORT RESULTS:")
    print("=" * 60)
    print(f"  ✅ Created (new):     {stats['created']}")
    print(f"  🔄 Merged (existing): {stats['merged']}")
    print(f"  ⏭️  Skipped (no phone): {stats['skipped_no_phone']}")
    print(f"  ❌ Errors:            {stats['errors']}")
    print(f"  📊 Total:             {sum(stats.values())}")
    if error_details:
        print(f"\nFirst 10 errors:")
        for e in error_details[:10]:
            print(f"  Row {e['row']}: {e['error']}")

asyncio.run(main())
