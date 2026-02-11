import asyncio
import sys
sys.path.insert(0, r'c:\Users\admin\kinyan-crm-new')

import openpyxl
from datetime import datetime
from db import SessionLocal
from db.models import Lead, Salesperson, Course
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError

# מיפוי אנשי מכירות (שם באקסל -> שם במערכת)
SALESPERSON_MAPPING = {
    "שרוליק": "ישראל ברים",
    "שלוימי גרוס": "שלמה גרוס",
    "אהרן מאירוביץ": "אהרן מאירוביץ",
    "משה גרינהויז": "משה גרינהויז",
    "נתנאל גפנר": "נתנאל גפנר",
    "שלמה דנציגר": "שלמה דנציגר",
    "N/A": None,
}

# מיפוי קורסים (שם באקסל -> שם במערכת)
COURSE_MAPPING = {
    "הלכות שבת": "שבת",
    "ממונות (חושן משפט)": None,  # אין במערכת
    "הלכות נידה/טהרה": "טהרה",
    "איסור והיתר": "איסור והיתר",
    "מסלול קניין שבת": "שבת",
    "השלים מבחן - טרם בחר הטבה": None,
    "מתעניין במסלול::": None,
}

# מיפוי סטטוסים (אקסל -> מערכת)
STATUS_MAPPING = {
    "ליד חדש": "ליד חדש",
    "ליד בתהליך": "ליד בתהליך",
    "חיוג ראשון": "ליד חדש",
    "לא רלוונטי": "לא רלוונטי",
    "ליד סגור - לקוח": "ליד סגור - לקוח",
}

# מיפוי סטטוס מענה (אקסל -> lead_response)
RESPONSE_MAPPING = {
    "נענה": "מעוניין",
    "ניתוק": "לא זמין",
    "לא נענה (Timeout)": "לא זמין",
    "פניה כללית": None,
}


def parse_date(date_str):
    """המרת תאריך מפורמט DD/MM/YYYY HH:MM:SS או DD/MM/YYYY HH:MM"""
    if not date_str:
        return None
    
    try:
        # ניסיון עם שניות
        return datetime.strptime(str(date_str), "%d/%m/%Y %H:%M:%S")
    except ValueError:
        try:
            # ניסיון בלי שניות
            return datetime.strptime(str(date_str), "%d/%m/%Y %H:%M")
        except ValueError:
            try:
                # ניסיון רק תאריך
                return datetime.strptime(str(date_str), "%d/%m/%Y")
            except ValueError:
                print(f"⚠️  לא הצלחתי לפרסר תאריך: {date_str}")
                return None


async def load_mappings(session):
    """טעינת מיפויים מהמערכת"""
    # טעינת אנשי מכירות
    salespeople_result = await session.execute(select(Salesperson))
    salespeople = {sp.name: sp.id for sp in salespeople_result.scalars()}
    
    # טעינת קורסים
    courses_result = await session.execute(select(Course))
    courses = {c.name: c.id for c in courses_result.scalars()}
    
    return salespeople, courses


async def import_leads(excel_path, limit=None, dry_run=False, quiet=False):
    """ייבוא לידים מאקסל"""
    
    # קריאת האקסל
    if not quiet:
        print("=" * 80)
        print("קורא קובץ Excel...")
        print("=" * 80)
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    
    total_rows = ws.max_row - 1
    if limit:
        total_rows = min(limit, total_rows)
    
    print(f"נמצאו {ws.max_row - 1} לידים בקובץ")
    if limit:
        print(f"מייבא רק {limit} לידים ראשונים")
    if dry_run:
        print("⚠️  מצב DRY RUN - לא ישמר כלום ל-DB")
    if quiet:
        print("מצב שקט - מציג התקדמות כל 50 לידים")
    print()
    
    async with SessionLocal() as session:
        # טעינת מיפויים
        salespeople_map, courses_map = await load_mappings(session)
        
        if not quiet:
            print("מיפויי אנשי מכירות:")
            for name, sp_id in salespeople_map.items():
                print(f"  {name} -> ID {sp_id}")
            
            print("\nמיפויי קורסים:")
            for name, c_id in courses_map.items():
                print(f"  {name} -> ID {c_id}")
            print()
        
        # סטטיסטיקות
        stats = {
            "success": 0,
            "skipped_duplicate": 0,
            "skipped_no_phone": 0,
            "errors": 0,
        }
        
        # עיבוד שורות
        print("=" * 80)
        print("מתחיל ייבוא...")
        print("=" * 80)
        
        for row_idx in range(2, min(2 + total_rows, ws.max_row + 1)):
            row_num = row_idx - 1
            
            # קריאת נתונים מהשורה
            row_data = {}
            for col_idx, header in enumerate(headers, 1):
                if header:
                    row_data[header] = ws.cell(row_idx, col_idx).value
            
            # טלפון הוא שדה חובה
            phone = row_data.get("טלפון ראשי")
            if not phone:
                if not quiet:
                    print(f"⏭️  שורה {row_num}: דילגתי - אין טלפון")
                stats["skipped_no_phone"] += 1
                continue
            
            phone = str(phone).strip()
            
            # בדיקת כפילות
            existing = await session.execute(
                select(Lead).where(Lead.phone == phone)
            )
            if existing.scalar_one_or_none():
                if not quiet:
                    print(f"⏭️  שורה {row_num}: דילגתי - טלפון {phone} כבר קיים")
                stats["skipped_duplicate"] += 1
                continue
            
            # בניית אובייקט Lead
            try:
                # שם מלא
                full_name = row_data.get("שם מלא") or "ליד ללא שם"
                
                # איש מכירות
                salesperson_name_excel = row_data.get("איש מכירות ")
                salesperson_id = None
                if salesperson_name_excel:
                    salesperson_name_excel = salesperson_name_excel.strip()
                    mapped_name = SALESPERSON_MAPPING.get(salesperson_name_excel)
                    if mapped_name:
                        salesperson_id = salespeople_map.get(mapped_name)
                
                # קורס
                course_name_excel = row_data.get("מוצר שמתעניין")
                course_id = None
                if course_name_excel:
                    course_name_excel = course_name_excel.strip()
                    mapped_course = COURSE_MAPPING.get(course_name_excel)
                    if mapped_course:
                        course_id = courses_map.get(mapped_course)
                
                # סטטוס
                status_excel = row_data.get("סטאטוס ליד") or "ליד חדש"
                status = STATUS_MAPPING.get(status_excel, "ליד חדש")
                
                # סטטוס מענה
                response_excel = row_data.get("סטטוס מענה")
                lead_response = None
                if response_excel:
                    lead_response = RESPONSE_MAPPING.get(response_excel.strip())
                
                # תאריכים
                arrival_date = parse_date(row_data.get("תאריך יצירה")) or datetime.now()
                last_contact = parse_date(row_data.get("תאריך פניה אחרונה"))
                
                # יצירת Lead
                lead = Lead(
                    full_name=full_name,
                    phone=phone,
                    email=row_data.get("מייל לקוח"),
                    city=row_data.get("עיר מגורים"),
                    address=row_data.get("כתובת"),
                    notes=row_data.get("הערות ליד"),
                    source_type="ייבוא ממערכת ישנה",
                    source_message=row_data.get("הודעה מהליד"),
                    source_name=row_data.get("שלוחת מוצר"),
                    campaign_name=row_data.get("שם המפרסם"),
                    arrival_date=arrival_date,
                    last_contact_date=last_contact,
                    status=status,
                    lead_response=lead_response,
                    salesperson_id=salesperson_id,
                    course_id=course_id,
                    created_by="import_script",
                )
                
                if not dry_run:
                    session.add(lead)
                    await session.flush()
                
                if not quiet:
                    print(f"✅ שורה {row_num}: {full_name} ({phone}) - {status}")
                stats["success"] += 1
                
                # הצגת התקדמות במצב שקט
                if quiet and stats["success"] % 50 == 0:
                    total_processed = sum(stats.values())
                    print(f"התקדמות: {total_processed}/{total_rows} | הצלחות: {stats['success']} | דילוגים: {stats['skipped_duplicate']}")
                
            except Exception as e:
                if not quiet:
                    print(f"❌ שורה {row_num}: שגיאה - {e}")
                stats["errors"] += 1
                continue
        
        # שמירה
        if not dry_run:
            await session.commit()
            print("\n💾 שמרתי את כל השינויים ל-DB")
        else:
            print("\n⚠️  DRY RUN - לא שמרתי כלום")
        
        # סיכום
        print("\n" + "=" * 80)
        print("סיכום ייבוא:")
        print("=" * 80)
        print(f"✅ הצלחות: {stats['success']}")
        print(f"⏭️  דילוגים (כפילויות): {stats['skipped_duplicate']}")
        print(f"⏭️  דילוגים (אין טלפון): {stats['skipped_no_phone']}")
        print(f"❌ שגיאות: {stats['errors']}")
        print(f"📊 סה\"כ: {sum(stats.values())}")
        print("=" * 80)


async def main():
    import argparse
    parser = argparse.ArgumentParser(description="ייבוא לידים מהמערכת הישנה")
    parser.add_argument("--limit", type=int, help="מספר לידים מקסימלי לייבוא")
    parser.add_argument("--dry-run", action="store_true", help="הרצה ללא שמירה ל-DB")
    parser.add_argument("--quiet", action="store_true", help="מצב שקט - פחות פלט")
    args = parser.parse_args()
    
    excel_path = r"C:\Users\admin\Downloads\לידים_02_11_2026.xlsx"
    await import_leads(excel_path, limit=args.limit, dry_run=args.dry_run, quiet=args.quiet)


if __name__ == "__main__":
    asyncio.run(main())
