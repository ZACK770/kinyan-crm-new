import asyncio
import sys
sys.path.insert(0, r'c:\Users\admin\kinyan-crm-new')

import openpyxl
from datetime import datetime
from db import SessionLocal
from db.models import Lead, Salesperson, Course
from sqlalchemy import select

# מיפוי אנשי מכירות
SALESPERSON_MAPPING = {
    "שרוליק": "ישראל ברים",
    "שלוימי גרוס": "שלמה גרוס",
    "אהרן מאירוביץ": "אהרן מאירוביץ",
    "משה גרינהויז": "משה גרינהויז",
    "נתנאל גפנר": "נתנאל גפנר",
    "שלמה דנציגר": "שלמה דנציגר",
    "N/A": None,
}

# מיפוי קורסים
COURSE_MAPPING = {
    "הלכות שבת": "שבת",
    "ממונות (חושן משפט)": None,
    "הלכות נידה/טהרה": "טהרה",
    "איסור והיתר": "איסור והיתר",
    "מסלול קניין שבת": "שבת",
    "השלים מבחן - טרם בחר הטבה": None,
    "מתעניין במסלול::": None,
}

# מיפוי סטטוסים
STATUS_MAPPING = {
    "ליד חדש": "ליד חדש",
    "ליד בתהליך": "ליד בתהליך",
    "חיוג ראשון": "ליד חדש",
    "לא רלוונטי": "לא רלוונטי",
    "ליד סגור - לקוח": "ליד סגור - לקוח",
}

# מיפוי סטטוס מענה
RESPONSE_MAPPING = {
    "נענה": "מעוניין",
    "ניתוק": "לא זמין",
    "לא נענה (Timeout)": "לא זמין",
    "פניה כללית": None,
}


def parse_date(date_str):
    if not date_str:
        return None
    try:
        return datetime.strptime(str(date_str), "%d/%m/%Y %H:%M:%S")
    except ValueError:
        try:
            return datetime.strptime(str(date_str), "%d/%m/%Y %H:%M")
        except ValueError:
            try:
                return datetime.strptime(str(date_str), "%d/%m/%Y")
            except ValueError:
                return None


async def main():
    print("מתחיל ייבוא...")
    
    # קריאת האקסל
    wb = openpyxl.load_workbook(r"C:\Users\admin\Downloads\לידים_02_11_2026.xlsx")
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    total_rows = ws.max_row - 1
    
    print(f"נמצאו {total_rows} לידים")
    
    async with SessionLocal() as session:
        # טעינת מיפויים
        salespeople_result = await session.execute(select(Salesperson))
        salespeople_map = {sp.name: sp.id for sp in salespeople_result.scalars()}
        
        courses_result = await session.execute(select(Course))
        courses_map = {c.name: c.id for c in courses_result.scalars()}
        
        # סטטיסטיקות
        stats = {"success": 0, "skipped_duplicate": 0, "skipped_no_phone": 0, "errors": 0}
        
        # עיבוד שורות
        for row_idx in range(2, ws.max_row + 1):
            row_num = row_idx - 1
            
            # התקדמות כל 100 לידים
            if row_num % 100 == 0:
                print(f"מעבד ליד {row_num}/{total_rows}...")
            
            row_data = {headers[col_idx]: ws.cell(row_idx, col_idx + 1).value 
                       for col_idx in range(len(headers)) if headers[col_idx]}
            
            phone = row_data.get("טלפון ראשי")
            if not phone:
                stats["skipped_no_phone"] += 1
                continue
            
            phone = str(phone).strip()
            
            # בדיקת כפילות
            existing = await session.execute(select(Lead).where(Lead.phone == phone))
            if existing.scalar_one_or_none():
                stats["skipped_duplicate"] += 1
                continue
            
            try:
                # בניית Lead
                full_name = row_data.get("שם מלא") or "ליד ללא שם"
                
                salesperson_name = row_data.get("איש מכירות ")
                salesperson_id = None
                if salesperson_name:
                    mapped = SALESPERSON_MAPPING.get(salesperson_name.strip())
                    if mapped:
                        salesperson_id = salespeople_map.get(mapped)
                
                course_name = row_data.get("מוצר שמתעניין")
                course_id = None
                if course_name:
                    mapped = COURSE_MAPPING.get(course_name.strip())
                    if mapped:
                        course_id = courses_map.get(mapped)
                
                status = STATUS_MAPPING.get(row_data.get("סטאטוס ליד", "ליד חדש"), "ליד חדש")
                
                response = row_data.get("סטטוס מענה")
                lead_response = RESPONSE_MAPPING.get(response.strip()) if response else None
                
                arrival_date = parse_date(row_data.get("תאריך יצירה")) or datetime.now()
                last_contact = parse_date(row_data.get("תאריך פניה אחרונה"))
                
                lead = Lead(
                    full_name=full_name,
                    phone=phone,
                    email=row_data.get("מייל לקוח"),
                    city=row_data.get("עיר מגורים"),
                    address=row_data.get("כתובת"),
                    notes=row_data.get("הערות ליד"),
                    source_type="ייבוא ממערכת ישנה",
                    source_message=row_data.get("הודעה מהליד"),
                    source_name=row_data.get("שלוחת מוצר"),
                    campaign_name=row_data.get("שם המפרסם"),
                    arrival_date=arrival_date,
                    last_contact_date=last_contact,
                    status=status,
                    lead_response=lead_response,
                    salesperson_id=salesperson_id,
                    course_id=course_id,
                    created_by="import_script",
                )
                
                session.add(lead)
                
                # commit כל 50 לידים
                if stats["success"] % 50 == 0 and stats["success"] > 0:
                    await session.commit()
                
                stats["success"] += 1
                
            except Exception as e:
                stats["errors"] += 1
                print(f"שגיאה בשורה {row_num}: {e}")
                continue
        
        # commit סופי
        await session.commit()
        
        # סיכום
        print("\n" + "=" * 60)
        print("סיכום ייבוא:")
        print("=" * 60)
        print(f"✅ הצלחות: {stats['success']}")
        print(f"⏭️  דילוגים (כפילויות): {stats['skipped_duplicate']}")
        print(f"⏭️  דילוגים (אין טלפון): {stats['skipped_no_phone']}")
        print(f"❌ שגיאות: {stats['errors']}")
        print(f"📊 סה\"כ: {sum(stats.values())}")
        print("=" * 60)


if __name__ == "__main__":
    asyncio.run(main())
