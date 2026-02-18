import openpyxl
from datetime import datetime

excel_path = r"C:\Users\admin\Downloads\לידים_02_11_2026.xlsx"

print("=" * 80)
print("ניתוח קובץ אקסל - בדיקת תאריכי יצירה")
print("=" * 80)

try:
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    
    print(f"\n✅ נמצאו {ws.max_row - 1} שורות באקסל")
    
    # הצגת כל העמודות
    print(f"\n📋 עמודות באקסל:")
    for i, h in enumerate(headers, 1):
        print(f"  {i}. {h}")
    
    # חיפוש עמודת תאריך יצירה
    date_col_idx = None
    date_col_name = None
    for i, h in enumerate(headers):
        if h and 'תאריך' in str(h) and ('יצירה' in str(h) or 'הגעה' in str(h)):
            date_col_idx = i + 1
            date_col_name = h
            break
    
    if date_col_idx:
        print(f"\n✅ נמצאה עמודת תאריך: '{date_col_name}' (עמודה {date_col_idx})")
        
        # בדיקת 10 דוגמאות
        print(f"\n📊 דוגמאות תאריכים (10 שורות ראשונות):")
        phone_col_idx = headers.index("טלפון ראשי") + 1 if "טלפון ראשי" in headers else None
        name_col_idx = headers.index("שם מלא") + 1 if "שם מלא" in headers else None
        
        dates_found = 0
        dates_empty = 0
        date_examples = []
        
        for row_idx in range(2, min(12, ws.max_row + 1)):
            phone = ws.cell(row_idx, phone_col_idx).value if phone_col_idx else "N/A"
            name = ws.cell(row_idx, name_col_idx).value if name_col_idx else "N/A"
            date_val = ws.cell(row_idx, date_col_idx).value
            
            if date_val:
                dates_found += 1
                date_type = type(date_val).__name__
                print(f"  שורה {row_idx-1}: {name} | {phone}")
                print(f"    תאריך: {date_val} (סוג: {date_type})")
                date_examples.append(date_val)
            else:
                dates_empty += 1
                print(f"  שורה {row_idx-1}: {name} | {phone}")
                print(f"    תאריך: ריק/None")
        
        # סטטיסטיקה כללית
        print(f"\n📊 סטטיסטיקה כללית:")
        total_with_dates = 0
        total_empty_dates = 0
        
        for row_idx in range(2, ws.max_row + 1):
            date_val = ws.cell(row_idx, date_col_idx).value
            if date_val:
                total_with_dates += 1
            else:
                total_empty_dates += 1
        
        print(f"  שורות עם תאריך: {total_with_dates}")
        print(f"  שורות ללא תאריך: {total_empty_dates}")
        
        if date_examples:
            print(f"\n📅 טווח תאריכים:")
            datetime_dates = []
            for d in date_examples:
                if isinstance(d, datetime):
                    datetime_dates.append(d)
            if datetime_dates:
                print(f"  מוקדם ביותר: {min(datetime_dates)}")
                print(f"  מאוחר ביותר: {max(datetime_dates)}")
    else:
        print(f"\n❌ לא נמצאה עמודת תאריך יצירה!")
        print("   עמודות תאריך זמינות:")
        for i, h in enumerate(headers, 1):
            if h and 'תאריך' in str(h):
                print(f"     {i}. {h}")

except FileNotFoundError:
    print(f"\n❌ קובץ לא נמצא: {excel_path}")
except Exception as e:
    print(f"\n❌ שגיאה: {e}")
    import traceback
    traceback.print_exc()

print("\n" + "=" * 80)
