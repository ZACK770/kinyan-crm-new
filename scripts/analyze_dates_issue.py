import asyncio
import sys
sys.path.insert(0, r'c:\Users\admin\kinyan-crm-new')

import openpyxl
from db import SessionLocal
from db.models import Lead
from sqlalchemy import select
from datetime import datetime

async def analyze():
    print("=" * 80)
    print("בדיקת בעיית תאריכי יצירה")
    print("=" * 80)
    
    # 1. קריאת האקסל
    excel_path = r"C:\Users\admin\Downloads\לידים_02_11_2026.xlsx"
    print(f"\n📂 קורא אקסל: {excel_path}")
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        
        print(f"✅ נמצאו {ws.max_row - 1} שורות באקסל")
        print(f"\nעמודות באקסל:")
        for i, h in enumerate(headers, 1):
            if h and 'תאריך' in str(h):
                print(f"  {i}. {h}")
        
        # בדיקת דוגמאות מהאקסל
        print(f"\n📊 דוגמאות תאריכים מהאקסל (5 שורות ראשונות):")
        date_col_idx = None
        for i, h in enumerate(headers):
            if h == "תאריך יצירה":
                date_col_idx = i + 1
                break
        
        if date_col_idx:
            for row_idx in range(2, min(7, ws.max_row + 1)):
                phone = ws.cell(row_idx, headers.index("טלפון ראשי") + 1).value
                date_val = ws.cell(row_idx, date_col_idx).value
                print(f"  שורה {row_idx-1}: טלפון={phone}, תאריך יצירה={date_val} (סוג: {type(date_val).__name__})")
        else:
            print("  ⚠️ לא נמצאה עמודה 'תאריך יצירה'")
    
    except FileNotFoundError:
        print(f"❌ קובץ לא נמצא: {excel_path}")
        print("   לא אוכל לקרוא נתונים מהאקסל")
    except Exception as e:
        print(f"❌ שגיאה בקריאת אקסל: {e}")
    
    # 2. בדיקת DB
    print(f"\n" + "=" * 80)
    print("בדיקת נתונים ב-DB")
    print("=" * 80)
    
    async with SessionLocal() as session:
        # ספירת לידים עם תאריך 11/2/2026
        result = await session.execute(
            select(Lead.id, Lead.full_name, Lead.phone, Lead.created_at, Lead.arrival_date, Lead.last_contact_date)
            .where(
                Lead.created_at >= '2026-02-11 00:00:00',
                Lead.created_at < '2026-02-12 00:00:00'
            )
            .limit(10)
        )
        leads_with_issue = result.all()
        
        if leads_with_issue:
            print(f"\n🔍 נמצאו לידים עם created_at = 11/2/2026:")
            print(f"   (מציג 10 ראשונים)")
            for lead in leads_with_issue:
                print(f"\n  📞 {lead.phone} - {lead.full_name}")
                print(f"     created_at: {lead.created_at}")
                print(f"     arrival_date: {lead.arrival_date}")
                print(f"     last_contact_date: {lead.last_contact_date}")
        
        # ספירה כללית
        from sqlalchemy import func
        count_result = await session.execute(
            select(func.count(Lead.id)).where(
                Lead.created_at >= '2026-02-11 00:00:00',
                Lead.created_at < '2026-02-12 00:00:00'
            )
        )
        total_count = count_result.scalar()
        print(f"\n📊 סה\"כ לידים עם created_at ב-11/2/2026: {total_count}")
        
        # בדיקה אם יש להם last_contact_date תקין
        valid_last_contact = await session.execute(
            select(func.count(Lead.id)).where(
                Lead.created_at >= '2026-02-11 00:00:00',
                Lead.created_at < '2026-02-12 00:00:00',
                Lead.last_contact_date.isnot(None),
                Lead.last_contact_date < '2026-02-11 00:00:00'
            )
        )
        valid_count = valid_last_contact.scalar()
        print(f"   מתוכם יש last_contact_date תקין (לפני 11/2): {valid_count}")
    
    print("\n" + "=" * 80)
    print("סיכום והמלצה")
    print("=" * 80)

if __name__ == "__main__":
    asyncio.run(analyze())
