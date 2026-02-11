import openpyxl
from collections import defaultdict

wb = openpyxl.load_workbook(r'C:\Users\admin\Downloads\לידים_02_11_2026.xlsx')
ws = wb.active

headers = [cell.value for cell in ws[1]]
print("=" * 80)
print("ניתוח קובץ לידים מהמערכת הישנה")
print("=" * 80)
print(f"\nסה\"כ לידים: {ws.max_row - 1}")
print(f"סה\"כ עמודות: {ws.max_column}")

# איסוף דוגמאות לכל עמודה
column_samples = defaultdict(list)
for row_idx in range(2, min(20, ws.max_row + 1)):
    for col_idx, header in enumerate(headers, 1):
        value = ws.cell(row_idx, col_idx).value
        if value is not None and header is not None:
            if len(column_samples[header]) < 5:
                column_samples[header].append(str(value))

# הדפסת דוגמאות לכל שדה
print("\n" + "=" * 80)
print("דוגמאות לכל שדה (עד 5 דוגמאות):")
print("=" * 80)

for header in headers:
    if header is None:
        continue
    non_empty = sum(1 for row in range(2, ws.max_row + 1) if ws.cell(row, headers.index(header) + 1).value is not None)
    fill_rate = (non_empty / (ws.max_row - 1)) * 100
    
    print(f"\n📌 {header}")
    print(f"   מולא: {non_empty}/{ws.max_row - 1} ({fill_rate:.1f}%)")
    
    if column_samples[header]:
        print(f"   דוגמאות:")
        for sample in column_samples[header]:
            print(f"      • {sample}")
    else:
        print(f"   (ריק)")

# ניתוח ערכים ייחודיים לשדות קטגוריים
print("\n" + "=" * 80)
print("ערכים ייחודיים לשדות קטגוריים:")
print("=" * 80)

categorical_fields = ['סטאטוס ליד', 'איש מכירות ', 'מוצר שמתעניין', 'סטטוס מענה', 'סוג תשלום']
for field in categorical_fields:
    if field not in headers:
        continue
    
    col_idx = headers.index(field) + 1
    values = set()
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row_idx, col_idx).value
        if val is not None:
            values.add(str(val))
    
    print(f"\n{field}:")
    if values:
        for val in sorted(values):
            count = sum(1 for row in range(2, ws.max_row + 1) if str(ws.cell(row, col_idx).value) == val)
            print(f"   • {val}: {count}")
    else:
        print(f"   (אין ערכים)")
