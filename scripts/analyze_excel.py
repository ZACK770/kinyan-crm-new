import openpyxl
import sys

# קריאת קובץ Excel
wb = openpyxl.load_workbook(r'C:\Users\admin\Downloads\לידים_02_11_2026.xlsx')
ws = wb.active

# הדפסת שמות העמודות
headers = [cell.value for cell in ws[1]]
print("Columns:", headers)
print(f"\nTotal rows: {ws.max_row}")
print(f"Total columns: {ws.max_column}")

# הדפסת 3 שורות ראשונות
print("\n=== First 3 data rows ===")
for row_idx in range(2, min(5, ws.max_row + 1)):
    row_data = {}
    for col_idx, header in enumerate(headers, 1):
        cell_value = ws.cell(row_idx, col_idx).value
        row_data[header] = cell_value
    print(f"\nRow {row_idx - 1}:")
    for k, v in row_data.items():
        if v is not None:
            print(f"  {k}: {v}")

# סטטיסטיקה על עמודות
print("\n=== Column Statistics ===")
for col_idx, header in enumerate(headers, 1):
    non_empty = sum(1 for row in range(2, ws.max_row + 1) if ws.cell(row, col_idx).value is not None)
    print(f"{header}: {non_empty}/{ws.max_row - 1} filled")
