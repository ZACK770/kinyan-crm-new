"""
Generic import API: dynamic entity discovery + field inference.
Supports any importable table without hardcoding entity names/fields.
"""
from fastapi import APIRouter, UploadFile, File, Query, Depends, Form, HTTPException
from datetime import datetime, timezone, date
from typing import Optional, Any, List, Dict
import io
import json

from sqlalchemy import select, or_, inspect, text, UniqueConstraint
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import DeclarativeBase
from sqlalchemy.sql.type_api import TypeEngine
import sqlalchemy

from db import get_db, Base
from utils.phone import normalize_phone
from .dependencies import require_permission

router = APIRouter()


def _is_importable_table(table_name: str, table_obj) -> bool:
    """Return True if the table is importable (has created_at or __importable__ flag)."""
    # Option A: has created_at column
    if any(col.name == "created_at" for col in table_obj.columns):
        return True
    # Option B: model has __importable__ = True
    model = Base.registry._class_registry.get(table_name)
    if getattr(model, "__importable__", False):
        return True
    return False


def _field_type_from_sqlalchemy(col_type: TypeEngine) -> str:
    """Map SQLAlchemy type to simple string for frontend."""
    try:
        if hasattr(col_type, "python_type"):
            py_type = col_type.python_type
        else:
            # fallback for exotic types
            return "string"
        
        # Handle common types safely
        if py_type is None:
            return "string"
        if issubclass(py_type, int):
            return "integer"
        if issubclass(py_type, float):
            return "numeric"
        if issubclass(py_type, bool):
            return "boolean"
        if issubclass(py_type, datetime):
            return "datetime"
        if issubclass(py_type, date):
            return "date"
        if issubclass(py_type, str):
            return "string"
        
        return "string"  # fallback for unknown types
    except Exception:
        return "string"  # ultimate fallback


def _infer_required_fields(table_obj) -> List[str]:
    """Suggest required fields: nullable=False except system/auto fields."""
    system_fields = {"id", "created_at", "updated_at"}
    required = []
    for col in table_obj.columns:
        if col.name in system_fields:
            continue
        if not col.nullable and col.default is None and col.server_default is None:
            required.append(col.name)
    return required


def _infer_duplicate_keys(table_obj) -> List[str]:
    """Suggest duplicate key candidates: unique columns or unique constraints."""
    candidates = []
    for col in table_obj.columns:
        if col.unique:
            candidates.append(col.name)
    for constraint in table_obj.constraints:
        if isinstance(constraint, UniqueConstraint):
            candidates.extend(constraint.columns.keys())
    return list(set(candidates))


def _get_entity_hebrew_name(table_name: str) -> str:
    """Return Hebrew name for common entities."""
    hebrew_names = {
        'leads': 'לידים',
        'students': 'סטודנטים', 
        'courses': 'קורסים',
        'payments': 'תשלומים',
        'commitments': 'התחייבויות',
        'campaigns': 'קמפיינים',
        'tasks': 'משימות',
        'users': 'משתמשים',
        'salespeople': 'נציגי מכירות',
        'lecturers': 'מרצים',
        'expenses': 'הוצאות',
        'collections': 'גבייה',
        'inquiries': 'פניות',
        'messages': 'הודעות',
        'audit_logs': 'יומן ביקורת',
        'webhook_logs': 'יומן וובהוקים',
    }
    return hebrew_names.get(table_name, table_name)


@router.get("/entities")
async def list_importable_entities(
    user=Depends(require_permission("admin")),
):
    """List all importable entities (tables)."""
    entities = []
    for table_name, table_obj in Base.metadata.tables.items():
        if _is_importable_table(table_name, table_obj):
            hebrew_name = _get_entity_hebrew_name(table_name)
            entities.append({
                "entity": table_name, 
                "label": f"{hebrew_name} ({table_name})"
            })
    return entities


@router.get("/entities/{entity}")
async def describe_entity_fields(
    entity: str,
    user=Depends(require_permission("admin")),
):
    """Return field schema for a given entity, including suggested required/duplicate fields."""
    if entity not in Base.metadata.tables:
        raise HTTPException(status_code=404, detail="Entity not found")
    table = Base.metadata.tables[entity]
    if not _is_importable_table(entity, table):
        raise HTTPException(status_code=403, detail="Entity not importable")

    fields = []
    for col in table.columns:
        is_identity = bool(getattr(col, "identity", None))
        is_computed = bool(getattr(col, "computed", None))
        fields.append({
            "name": col.name,
            "type": _field_type_from_sqlalchemy(col.type),
            "nullable": col.nullable,
            "primary_key": col.primary_key,
            "unique": col.unique,
            "foreign_key": bool(col.foreign_keys),
            "writable": (
                not col.primary_key
                and col.autoincrement is not True
                and not is_identity
                and not is_computed
            ),
        })

    return {
        "entity": entity,
        "fields": fields,
        "required_fields_suggested": _infer_required_fields(table),
        "duplicate_keys_suggested": _infer_duplicate_keys(table),
    }


def _describe_entity_fields_sync(entity: str) -> dict:
    """Synchronous version of describe_entity_fields for internal use."""
    if entity not in Base.metadata.tables:
        raise HTTPException(status_code=404, detail="Entity not found")
    table = Base.metadata.tables[entity]
    if not _is_importable_table(entity, table):
        raise HTTPException(status_code=403, detail="Entity not importable")

    fields = []
    for col in table.columns:
        is_identity = bool(getattr(col, "identity", None))
        is_computed = bool(getattr(col, "computed", None))
        fields.append({
            "name": col.name,
            "type": _field_type_from_sqlalchemy(col.type),
            "nullable": col.nullable,
            "primary_key": col.primary_key,
            "unique": col.unique,
            "foreign_key": bool(col.foreign_keys),
            "writable": (
                not col.primary_key
                and col.autoincrement is not True
                and not is_identity
                and not is_computed
            ),
        })

    return {
        "entity": entity,
        "fields": fields,
        "required_fields_suggested": _infer_required_fields(table),
        "duplicate_keys_suggested": _infer_duplicate_keys(table),
    }


@router.post("/preview-file")
async def preview_import_file(
    entity: str = Form(...),
    file: UploadFile = File(...),
    user=Depends(require_permission("admin")),
):
    """Return headers + sample rows for mapping UI for a given entity."""
    try:
        if entity not in Base.metadata.tables:
            raise HTTPException(status_code=404, detail="Entity not found")
        table = Base.metadata.tables[entity]
        if not _is_importable_table(entity, table):
            raise HTTPException(status_code=403, detail="Entity not importable")

        import openpyxl
        content = await file.read()
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            sheet = wb.active
            rows_iter = sheet.iter_rows(values_only=True)
            headers_raw = next(rows_iter, [])
            headers = [str(h) if h is not None else "" for h in headers_raw]
            sample_rows = []
            for i, row in enumerate(rows_iter):
                if i >= 5:
                    break
                sample_rows.append([str(cell) if cell is not None else "" for cell in row])
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {e}")

        # Include field schema from describe endpoint
        fields_info = _describe_entity_fields_sync(entity)

        return {
            "entity": entity,
            "headers": headers,
            "sample_rows": sample_rows,
            "fields": fields_info["fields"],
            "required_fields_suggested": fields_info["required_fields_suggested"],
            "duplicate_keys_suggested": fields_info["duplicate_keys_suggested"],
        }
    except HTTPException:
        raise  # re-raise HTTP exceptions as-is
    except Exception as e:
        import logging
        logging.error(f"Preview file error: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")


async def _find_existing_generic(
    db: AsyncSession,
    entity: str,
    duplicate_key_field: str,
    value: Any,
) -> Optional[Any]:
    """Find existing row by duplicate_key_field for any entity."""
    if entity not in Base.metadata.tables:
        return None
    table = Base.metadata.tables[entity]
    col = table.columns.get(duplicate_key_field)
    if col is None:
        return None
    # Special case for phone fields: apply normalize logic
    if "phone" in duplicate_key_field.lower() and isinstance(value, str):
        norm = normalize_phone(value)
        if not norm:
            return None
        # Try exact match, then fallback without leading 0
        stmt = select(table).where(col == norm).limit(1)
        result = await db.execute(stmt)
        row = result.scalar_one_or_none()
        if row:
            return row
        if norm.startswith("0") and len(norm) >= 2:
            short = norm[1:]
            stmt = select(table).where(or_(col == short, col.contains(short))).limit(1)
            result = await db.execute(stmt)
            return result.scalar_one_or_none()
        return None
    else:
        stmt = select(table).where(col == value).limit(1)
        result = await db.execute(stmt)
        return result.scalar_one_or_none()


@router.post("/import")
async def import_generic_entity(
    entity: str = Form(...),
    duplicate_mode: str = Form("skip"),
    duplicate_key_field: Optional[str] = Form(None),
    required_fields_override: Optional[str] = Form(None),  # JSON string list
    file: UploadFile = File(...),
    mapping_json: str = Form(...),
    user=Depends(require_permission("admin")),
    db: AsyncSession = Depends(get_db),
):
    """Generic import for any entity."""
    if entity not in Base.metadata.tables:
        raise HTTPException(status_code=404, detail="Entity not found")
    table = Base.metadata.tables[entity]
    if not _is_importable_table(entity, table):
        raise HTTPException(status_code=403, detail="Entity not importable")

    mapping = json.loads(mapping_json)
    required_fields = (
        json.loads(required_fields_override)
        if required_fields_override
        else _infer_required_fields(table)
    )

    # Load Excel
    import openpyxl
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        sheet = wb.active
        rows_iter = sheet.iter_rows(values_only=True)
        headers_raw = next(rows_iter, [])
        headers = [str(h) if h is not None else "" for h in headers_raw]
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {e}")

    stats = {
        "total_rows": 0,
        "created": 0,
        "updated": 0,
        "skipped_dup": 0,
        "errors": 0,
    }
    error_details = []

    for row_idx, raw_cells in enumerate(rows_iter, start=2):
        stats["total_rows"] += 1
        try:
            row_dict = {headers[i]: raw_cells[i] for i in range(min(len(headers), len(raw_cells)))}
            # Build data dict from mapping
            data = {}
            for field, excel_header in mapping.items():
                val = row_dict.get(excel_header)
                if val is not None and val != "":
                    # Basic type conversion (could be extended)
                    col = table.columns.get(field)
                    if col is not None:
                        py_type = _field_type_from_sqlalchemy(col.type)
                        try:
                            if py_type == "integer":
                                data[field] = int(float(val))
                            elif py_type == "numeric":
                                data[field] = float(val)
                            elif py_type == "boolean":
                                data[field] = str(val).strip().lower() in {"true", "1", "yes", "כן"}
                            elif py_type in {"datetime", "date"}:
                                # Try parsing common formats
                                if isinstance(val, (datetime,)):
                                    data[field] = val
                                else:
                                    data[field] = datetime.strptime(str(val), "%Y-%m-%d")
                            else:
                                data[field] = str(val).strip()
                        except Exception:
                            # Keep as string if conversion fails
                            data[field] = str(val).strip()
                    else:
                        data[field] = str(val).strip()

            # Validate required fields
            missing = [f for f in required_fields if f not in data or data[f] in (None, "")]
            if missing:
                stats["errors"] += 1
                error_details.append({"row": row_idx - 1, "error": f"Missing required: {', '.join(missing)}"})
                continue

            # System fields
            if "created_at" in table.columns:
                data["created_at"] = datetime.now(timezone.utc)
            if "created_by" in table.columns:
                data["created_by"] = f"import_excel:{getattr(user, 'email', 'admin')}"

            # Duplicate detection
            existing = None
            if duplicate_key_field and duplicate_key_field in data:
                existing = await _find_existing_generic(db, entity, duplicate_key_field, data[duplicate_key_field])

            if existing:
                if duplicate_mode == "skip":
                    stats["skipped_dup"] += 1
                    continue
                elif duplicate_mode == "merge":
                    # Update only empty fields
                    for k, v in data.items():
                        if k != duplicate_key_field and v is not None:
                            cur = getattr(existing, k, None)
                            if cur in (None, ""):
                                setattr(existing, k, v)
                    await db.commit()
                    stats["updated"] += 1
                    continue
                elif duplicate_mode == "overwrite":
                    # Overwrite all fields
                    for k, v in data.items():
                        if v is not None:
                            setattr(existing, k, v)
                    await db.commit()
                    stats["updated"] += 1
                    continue
                else:
                    # Unknown mode: treat as skip
                    stats["skipped_dup"] += 1
                    continue

            # Insert new
            try:
                new_row = table.insert().values(**data)
                await db.execute(new_row)
                await db.commit()
                stats["created"] += 1
            except Exception as e:
                await db.rollback()
                stats["errors"] += 1
                error_details.append({"row": row_idx - 1, "error": str(e)})
                continue
        except Exception as e:
            try:
                await db.rollback()
            except Exception:
                pass
            stats["errors"] += 1
            error_details.append({"row": row_idx - 1, "error": str(e)})
            continue

    return {
        "message": f"Import completed for {entity}",
        "entity": entity,
        "total_rows": stats["total_rows"],
        "stats": stats,
        "errors": error_details,
    }
