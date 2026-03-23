""" 
API endpoint for importing leads from Excel files.
Temporary utility for migrating data from the old system.
"""
from fastapi import APIRouter, UploadFile, File, Query, Depends, Form, HTTPException
from datetime import datetime, timezone
from typing import Optional, Any
import io
import json

from sqlalchemy import select, or_
from sqlalchemy.ext.asyncio import AsyncSession

from db import get_db
from db.models import Lead, Salesperson, Course
from utils.phone import normalize_phone
from .dependencies import require_permission

router = APIRouter()


ALLOWED_LEAD_IMPORT_FIELDS: set[str] = {
    "full_name",
    "family_name",
    "phone",
    "phone2",
    "email",
    "city",
    "address",
    "id_number",
    "notes",
    "source_type",
    "source_message",
    "campaign_name",
    "requested_course",
    "status",
    "lead_response",
    "salesperson_name",
    "course_name",
    "created_at",
    "arrival_date",
    "last_contact_date",
}

# מיפוי אנשי מכירות (שם באקסל -> שם במערכת)
SALESPERSON_MAPPING = {
    "שרוליק": "ישראל ברים",
    "שלוימי גרוס": "שלמה גרוס",
    "אהרן מאירוביץ": "אהרן מאירוביץ",
    "משה גרינהויז": "משה גרינהויז",
    "נתנאל גפנר": "נתנאל גפנר",
    "שלמה דנציגר": "שלמה דנציגר",
    "אברימי ברים": "אברימי ברים",
    "דודי וצלר": "דודי וצלר",
    "נפתלי לרנר": "נפתלי לרנר",
    "מוטי העכט": "מוטי העכט",
    "חיים ברים": "חיים ברים",
    "מרדכי ארנפלד": "מרדכי ארנפלד",
    "מוטי דבלינגר": "מוטי דבלינגר",
    "N/A": None,
}

# מיפוי קורסים (שם באקסל -> שם במערכת)
COURSE_MAPPING = {
    "הלכות שבת": "שבת",
    "ממונות (חושן משפט)": None,
    "הלכות נידה/טהרה": "טהרה",
    "איסור והיתר": "איסור והיתר",
    "מסלול קניין שבת": "שבת",
    "השלים מבחן - טרם בחר הטבה": None,
    "מתעניין במסלול::": None,
}

# מיפוי סטטוסים (אקסל -> מערכת)
STATUS_MAPPING = {
    "ליד חדש": "ליד חדש",
    "ליד בתהליך": "ליד בתהליך",
    "חיוג ראשון": "חיוג ראשון",
    "נסלק": "נסלק",
    "תלמיד פעיל": "תלמיד פעיל",
    "לא רלוונטי": "לא רלוונטי",
    # Legacy mappings — old values → new values
    "במעקב": "ליד בתהליך",
    "מתעניין": "ליד בתהליך",
    "ליד סגור - לקוח": "נסלק",
    "ליד סגור - לא רלוונטי": "לא רלוונטי",
    "converted": "נסלק",
}

# מיפוי סטטוס מענה
RESPONSE_MAPPING = {
    "נענה": "מעוניין",
    "ניתוק": "לא זמין",
    "לא נענה (Timeout)": "לא זמין",
    "פניה כללית": None,
}


def _get_cell(row: dict, *keys):
    """מחזיר את הערך הראשון שנמצא מתוך מספר שמות אפשריים לעמודה"""
    for k in keys:
        val = row.get(k)
        if val is not None and str(val).strip():
            return val
    return None


def _safe_str(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


async def _load_lookup_maps(db: AsyncSession) -> tuple[dict[str, int], dict[str, int]]:
    sp_res = await db.execute(select(Salesperson))
    sp_ids = {sp.name: sp.id for sp in sp_res.scalars()}
    c_res = await db.execute(select(Course))
    c_ids = {c.name: c.id for c in c_res.scalars()}
    return sp_ids, c_ids


def _row_from_headers_and_cells(headers: list[Any], cells: list[Any]) -> dict[str, Any]:
    row: dict[str, Any] = {}
    for i in range(min(len(headers), len(cells))):
        h = headers[i]
        if h is None:
            continue
        key = str(h).strip()
        if not key:
            continue
        row[key] = cells[i]
    return row


async def _find_existing_lead_by_phone(db: AsyncSession, phone: str) -> Lead | None:
    """Find existing lead by phone with tolerance for leading zero differences."""
    clean = normalize_phone(phone)
    if not clean:
        return None

    # Exact match
    res = await db.execute(select(Lead).where(Lead.phone == clean).limit(1))
    lead = res.scalar_one_or_none()
    if lead:
        return lead

    # If stored without leading 0 (legacy), try to match by the short version.
    if clean.startswith("0") and len(clean) >= 2:
        short = clean[1:]
        res = await db.execute(
            select(Lead)
            .where(or_(Lead.phone == short, Lead.phone.contains(short)))
            .limit(1)
        )
        return res.scalar_one_or_none()

    return None


@router.post("/import-leads/preview-file")
async def preview_import_leads_file(
    file: UploadFile = File(...),
    user=Depends(require_permission("admin")),
):
    """Return headers + a small sample of rows for the mapping UI."""
    import openpyxl

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"קובץ אקסל לא תקין: {e}")

    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    clean_headers = [str(h).strip() for h in headers if h is not None and str(h).strip()]

    sample_rows: list[dict[str, Any]] = []
    for row_idx in range(2, min(ws.max_row + 1, 7)):
        cells = [ws.cell(row_idx, col_idx + 1).value for col_idx in range(len(headers))]
        row = _row_from_headers_and_cells(headers, cells)
        # stringify values for UI
        sample_rows.append({k: _safe_str(v) for k, v in row.items()})

    return {
        "headers": clean_headers,
        "sample_rows": sample_rows,
        "supported_fields": sorted(ALLOWED_LEAD_IMPORT_FIELDS),
        "required_fields": ["full_name", "phone"],
    }


def parse_date(d):
    """המרת תאריך — תומך ב-datetime objects מ-openpyxl וגם בפורמטי טקסט"""
    if not d:
        return None
    # openpyxl מחזיר datetime object ישירות אם התא מפורמט כתאריך
    if isinstance(d, datetime):
        if d.tzinfo is None:
            return d.replace(tzinfo=timezone.utc)
        return d
    # ניסיון פרסור מטקסט
    text = str(d).strip()
    for fmt in ["%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y",
                "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d",
                "%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M", "%d.%m.%Y"]:
        try:
            dt = datetime.strptime(text, fmt)
            return dt.replace(tzinfo=timezone.utc)
        except ValueError:
            pass
    return None


@router.post("/import-leads")
async def import_leads_from_excel(
    file: UploadFile = File(...),
    duplicate_mode: str = Query(default="skip", description="skip / merge / overwrite / update_field"),
    update_field_name: Optional[str] = Query(default=None, description="שם השדה לעדכון במצב update_field"),
    mapping_json: Optional[str] = Form(default=None, description="JSON mapping: lead_field -> excel_header"),
    user=Depends(require_permission("admin")),
    db: AsyncSession = Depends(get_db),
):
    """
    ייבוא לידים מקובץ אקסל.
    
    duplicate_mode:
    - skip: דילוג על לידים עם טלפון קיים
    - merge: מיזוג - עדכון שדות ריקים בלבד (לא דורס נתונים קיימים)
    - overwrite: דריסה מלאה של הליד הקיים
    - update_field: עדכון שדה ספציפי בלבד (צריך לציין update_field_name)
    """
    import openpyxl

    # קריאת הקובץ
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    total_rows = ws.max_row - 1

    stats = {"created": 0, "merged": 0, "overwritten": 0, "updated": 0, "skipped_dup": 0, "skipped_no_phone": 0, "skipped_not_found": 0, "errors": 0}
    error_details = []
    
    # במצב update_field - חייב לציין את שם השדה
    if duplicate_mode == "update_field" and not update_field_name:
        return {"message": "שגיאה: במצב update_field חובה לציין את update_field_name", "stats": stats}

    # mapping_json is optional for backwards compatibility.
    mapping: dict[str, str] | None = None
    if mapping_json:
        try:
            mapping_raw = json.loads(mapping_json)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"mapping_json לא תקין: {e}")
        if not isinstance(mapping_raw, dict):
            raise HTTPException(status_code=400, detail="mapping_json חייב להיות אובייקט")
        mapping = {}
        for k, v in mapping_raw.items():
            if not isinstance(k, str) or not isinstance(v, str):
                continue
            key = k.strip()
            val = v.strip()
            if not key or not val:
                continue
            if key not in ALLOWED_LEAD_IMPORT_FIELDS:
                raise HTTPException(status_code=400, detail=f"שדה לא נתמך במיפוי: {key}")
            mapping[key] = val

        missing = [f for f in ("full_name", "phone") if f not in mapping]
        if missing:
            raise HTTPException(status_code=400, detail=f"חסר מיפוי לשדות חובה: {', '.join(missing)}")

    sp_ids, c_ids = await _load_lookup_maps(db)

    for row_idx in range(2, ws.max_row + 1):
        cells = [ws.cell(row_idx, col_idx + 1).value for col_idx in range(len(headers))]
        row = _row_from_headers_and_cells(headers, cells)

        if mapping:
            phone_raw = row.get(mapping.get("phone", ""))
        else:
            phone_raw = row.get("טלפון ראשי")
        phone = normalize_phone(str(phone_raw)) if phone_raw else ""
        if not phone:
            stats["skipped_no_phone"] += 1
            continue

        # חישוב שדות
        try:
            def col(field: str, *fallback_keys: str):
                if mapping and field in mapping:
                    return row.get(mapping[field])
                return _get_cell(row, *fallback_keys)

            full_name = _safe_str(col("full_name", "שם מלא")) or "ליד ללא שם"
            family_name = _safe_str(col("family_name", "משפחה"))

            # Salesperson (mapped by name)
            sp_name_raw = _safe_str(col("salesperson_name", "איש מכירות", "איש מכירות "))
            sp_id = None
            if sp_name_raw:
                sp_mapped = SALESPERSON_MAPPING.get(sp_name_raw.strip(), sp_name_raw.strip())
                sp_id = sp_ids.get(sp_mapped) if sp_mapped else None

            requested_course = _safe_str(col("requested_course", "מוצר שמתעניין", "מוצר", "שם קורס", "קורס"))

            # Course (mapped by name)
            course_name_raw = _safe_str(col("course_name", "קורס"))
            course_name = course_name_raw or requested_course
            c_id = None
            if course_name:
                mapped_course = COURSE_MAPPING.get(course_name.strip(), course_name.strip())
                c_id = c_ids.get(mapped_course) if mapped_course else None

            resp = _safe_str(col("lead_response", "סטטוס מענה"))
            lead_response = RESPONSE_MAPPING.get(resp.strip()) if resp else None

            status_raw = _safe_str(col("status", "סטאטוס ליד")) or "ליד חדש"
            status = STATUS_MAPPING.get(status_raw.strip(), "ליד חדש")

            created_date = parse_date(col("created_at", "תאריך יצירה", "תאריך הגעה")) or datetime.now(timezone.utc)
            arrival_date = parse_date(col("arrival_date", "תאריך הגעה")) or created_date
            last_contact = parse_date(col("last_contact_date", "תאריך פניה אחרונה"))

            phone2_raw = col("phone2", "טלפון נוסף")
            phone2 = normalize_phone(str(phone2_raw)) if phone2_raw else None

            id_number = _safe_str(col("id_number", "תעודת זהות"))

            notes = _safe_str(col("notes", "הערות ליד", "הערות על הליד"))

            email = _safe_str(col("email", "מייל לקוח"))
            city = _safe_str(col("city", "עיר מגורים"))
            address = _safe_str(col("address", "כתובת"))
            source_message = _safe_str(col("source_message", "הודעה מהליד"))

            source_type = _safe_str(col("source_type", "מקור הגעה כללי")) or "ייבוא מאקסל"
            campaign_name = _safe_str(col("campaign_name", "שם המפרסם", "שם קמפיין"))

            new_data = dict(
                full_name=full_name,
                family_name=family_name,
                phone=phone,
                phone2=phone2,
                email=email,
                city=city,
                address=address,
                id_number=id_number,
                notes=notes,
                source_type=source_type,
                source_message=source_message,
                campaign_name=campaign_name,
                requested_course=requested_course,
                arrival_date=arrival_date,
                last_contact_date=last_contact,
                status=status,
                lead_response=lead_response,
                salesperson_id=sp_id,
                course_id=c_id,
                created_at=created_date,
                created_by=f"import_excel:{user.email if getattr(user, 'email', None) else 'admin'}",
            )
        except Exception as e:
            stats["errors"] += 1
            error_details.append({"row": row_idx - 1, "error": str(e)})
            continue

        # בדיקת כפילות
        existing = await _find_existing_lead_by_phone(db, phone)

        if existing:
            if duplicate_mode == "skip":
                stats["skipped_dup"] += 1
                continue
            elif duplicate_mode == "merge":
                # מיזוג - עדכון שדות ריקים בלבד
                for key, val in new_data.items():
                    if val is not None and key != "phone":
                        current = getattr(existing, key, None)
                        if current is None or current == "" or current == "ליד ללא שם":
                            setattr(existing, key, val)
                # created_at — תמיד מעדכן מהאקסל (תאריך יצירה אמיתי)
                if new_data.get("created_at"):
                    existing.created_at = new_data["created_at"]
                    existing.arrival_date = new_data["created_at"]
                # הערות - צירוף
                if new_data.get("notes") and existing.notes:
                    if new_data["notes"] not in existing.notes:
                        existing.notes = existing.notes + "\n---\n" + new_data["notes"]
                stats["merged"] += 1
            elif duplicate_mode == "overwrite":
                # דריסה מלאה
                for key, val in new_data.items():
                    if key != "phone":
                        setattr(existing, key, val)
                stats["overwritten"] += 1
            elif duplicate_mode == "update_field":
                # עדכון שדה ספציפי בלבד
                if update_field_name in new_data:
                    val = new_data[update_field_name]
                    if val is not None:
                        setattr(existing, update_field_name, val)
                        stats["updated"] += 1
                    else:
                        stats["skipped_dup"] += 1
                else:
                    stats["errors"] += 1
                    error_details.append({"row": row_idx - 1, "error": f"שדה {update_field_name} לא נמצא בנתונים"})
        else:
            # במצב update_field - אם הליד לא קיים, דלג
            if duplicate_mode == "update_field":
                stats["skipped_not_found"] += 1
            else:
                # ליד חדש
                db.add(Lead(**new_data))
                stats["created"] += 1

    await db.commit()

    return {
        "message": "ייבוא הושלם!",
        "total_rows": total_rows,
        "stats": stats,
        "errors": error_details[:20],  # מקסימום 20 שגיאות
    }


@router.get("/import-leads/preview")
async def preview_excel_columns():
    """מחזיר את רשימת העמודות הנתמכות לייבוא"""
    return {
        "supported_columns": [
            {"excel": "שם מלא", "field": "full_name", "required": True},
            {"excel": "טלפון ראשי", "field": "phone", "required": True},
            {"excel": "מייל לקוח", "field": "email", "required": False},
            {"excel": "עיר מגורים", "field": "city", "required": False},
            {"excel": "כתובת", "field": "address", "required": False},
            {"excel": "הערות ליד", "field": "notes", "required": False},
            {"excel": "איש מכירות ", "field": "salesperson_id", "required": False},
            {"excel": "מוצר שמתעניין", "field": "course_id", "required": False},
            {"excel": "סטאטוס ליד", "field": "status", "required": False},
            {"excel": "סטטוס מענה", "field": "lead_response", "required": False},
            {"excel": "הודעה מהליד", "field": "source_message", "required": False},
            {"excel": "שם המפרסם", "field": "campaign_name", "required": False},
            {"excel": "תאריך יצירה", "field": "arrival_date", "required": False},
            {"excel": "תאריך פניה אחרונה", "field": "last_contact_date", "required": False},
        ],
        "duplicate_modes": [
            {"value": "skip", "label": "דילוג", "description": "דילוג על לידים עם טלפון קיים"},
            {"value": "merge", "label": "מיזוג", "description": "עדכון שדות ריקים בלבד (לא דורס נתונים קיימים)"},
            {"value": "overwrite", "label": "דריסה", "description": "דריסה מלאה של הליד הקיים"},
        ],
        "status_mapping": STATUS_MAPPING,
        "salesperson_mapping": SALESPERSON_MAPPING,
        "course_mapping": COURSE_MAPPING,
    }
