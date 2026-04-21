"""
Kinyan CRM - FastAPI Application Entry Point
Minimal app.py - all logic lives in services/
Serves both API and Frontend from a single server!
"""
# Load .env FIRST, before any other imports that use environment variables
from dotenv import load_dotenv
load_dotenv()

import os
from pathlib import Path
from contextlib import asynccontextmanager
from fastapi import FastAPI, Request, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from starlette.exceptions import HTTPException as StarletteHTTPException

from db import init_db
from api import test_netfree_api, leads_api, students_api, courses_api, dashboard_api, webhooks_api
from api import inquiries_api, exams_api, payments_api, expenses_api, attendance_api, collections_api
from api import auth_api, users_api, audit_logs_api, campaigns_api, files_api, sales_assignment_api
from api import course_tracks_api, lecturers_api, messages_api, templates_api, lead_conversion_api, chat_api
from api import salespeople_api, tasks_api, examinees_api, table_prefs_api, popup_api, webhook_logs_api
from api import exam_registration_api, export_api, import_api, import_generic_api, topics_api
from api import deliveries_api

# Frontend build directory
FRONTEND_DIR = Path(__file__).parent / "frontend" / "dist"


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Startup: create tables if needed."""
    await init_db()
    yield


app = FastAPI(
    title="Kinyan CRM",
    description="מערכת CRM לניהול שיווק ולמידה - קניין הוראה",
    version="1.0.0",
    lifespan=lifespan,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- API Routes ---
app.include_router(test_netfree_api.router)
app.include_router(leads_api.router, prefix="/api/leads", tags=["leads"])
app.include_router(lead_conversion_api.router)  # Lead conversion endpoints (uses /api/leads prefix)
app.include_router(students_api.router, prefix="/api/students", tags=["students"])
app.include_router(courses_api.router, prefix="/api/courses", tags=["courses"])
app.include_router(course_tracks_api.router)
app.include_router(lecturers_api.router, prefix="/api/lecturers", tags=["lecturers"])
app.include_router(exams_api.router, prefix="/api/exams", tags=["exams"])
app.include_router(attendance_api.router, prefix="/api/attendance", tags=["attendance"])
app.include_router(payments_api.router, prefix="/api/finance", tags=["finance"])
app.include_router(collections_api.router, prefix="/api/collections", tags=["collections"])
app.include_router(expenses_api.router, prefix="/api/expenses", tags=["expenses"])
app.include_router(inquiries_api.router, prefix="/api/inquiries", tags=["inquiries"])
app.include_router(dashboard_api.router, prefix="/api/dashboard", tags=["dashboard"])
app.include_router(campaigns_api.router, prefix="/api/campaigns", tags=["campaigns"])
app.include_router(webhooks_api.router, prefix="/webhooks", tags=["webhooks"])

# --- Auth & User Management ---
app.include_router(auth_api.router, prefix="/api/auth", tags=["auth"])
app.include_router(users_api.router, prefix="/api/users", tags=["users"])

# --- System Management ---
app.include_router(audit_logs_api.router, prefix="/api/audit-logs", tags=["audit-logs"])
app.include_router(files_api.router, prefix="/api/files", tags=["files"])
app.include_router(sales_assignment_api.router, prefix="/api/sales-assignment-rules", tags=["sales-assignment"])
app.include_router(messages_api.router, prefix="/api/messages", tags=["messages"])
app.include_router(templates_api.router, prefix="/api/templates", tags=["templates"])
app.include_router(chat_api.router, prefix="/api/chat", tags=["chat"])
app.include_router(salespeople_api.router, prefix="/api/salespeople", tags=["salespeople"])
app.include_router(tasks_api.router, prefix="/api/tasks", tags=["tasks"])
app.include_router(examinees_api.router, prefix="/api/examinees", tags=["examinees"])
app.include_router(table_prefs_api.router, prefix="/api/table-prefs", tags=["table-prefs"])
app.include_router(popup_api.router, prefix="/api/popups", tags=["popups"])
app.include_router(webhook_logs_api.router, prefix="/api/webhook-logs", tags=["webhook-logs"])
app.include_router(exam_registration_api.router, prefix="/api/exam-registration", tags=["exam-registration"])
app.include_router(export_api.router, prefix="/api/export", tags=["export"])
app.include_router(import_api.router, prefix="/api/import", tags=["import"])
app.include_router(import_generic_api.router, prefix="/api/admin/import", tags=["import-generic"])
app.include_router(topics_api.router, prefix="/api/topics", tags=["topics"])


# Health check endpoint (for Render and monitoring)
@app.get("/health")
async def health():
    return {"status": "healthy"}


# API status endpoint
@app.get("/api/status")
async def api_status():
    return {"status": "ok", "app": "Kinyan CRM", "version": "1.0.0"}


# --- TEMPORARY: Import old leads from Excel ---
@app.get("/api/import-old-leads")
async def import_old_leads():
    import openpyxl
    from datetime import datetime as dt
    from db import SessionLocal
    from db.models import Lead, Salesperson, Course
    from sqlalchemy import select

    SP_MAP = {"שרוליק": "ישראל ברים", "שלוימי גרוס": "שלמה גרוס", "אהרן מאירוביץ": "אהרן מאירוביץ", "משה גרינהויז": "משה גרינהויז", "נתנאל גפנר": "נתנאל גפנר", "שלמה דנציגר": "שלמה דנציגר", "N/A": None}
    C_MAP = {"הלכות שבת": "שבת", "ממונות (חושן משפט)": None, "הלכות נידה/טהרה": "טהרה", "איסור והיתר": "איסור והיתר", "מסלול קניין שבת": "שבת", "השלים מבחן - טרם בחר הטבה": None, "מתעניין במסלול::": None}
    ST_MAP = {"ליד חדש": "ליד חדש", "ליד בתהליך": "ליד בתהליך", "חיוג ראשון": "ליד חדש", "לא רלוונטי": "לא רלוונטי", "ליד סגור - לקוח": "ליד סגור - לקוח"}
    R_MAP = {"נענה": "מעוניין", "ניתוק": "לא זמין", "לא נענה (Timeout)": "לא זמין", "פניה כללית": None}

    def pd(d):
        if not d: return None
        for f in ["%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y"]:
            try: return dt.strptime(str(d), f)
            except: pass
        return None

    wb = openpyxl.load_workbook(r"C:\Users\admin\Downloads\לידים_02_11_2026.xlsx")
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    stats = {"ok": 0, "dup": 0, "skip": 0, "err": 0}

    async with SessionLocal() as session:
        sp_res = await session.execute(select(Salesperson))
        sp_ids = {sp.name: sp.id for sp in sp_res.scalars()}
        c_res = await session.execute(select(Course))
        c_ids = {c.name: c.id for c in c_res.scalars()}

        for row_idx in range(2, ws.max_row + 1):
            row = {headers[i]: ws.cell(row_idx, i+1).value for i in range(len(headers)) if headers[i]}
            phone = row.get("טלפון ראשי")
            if not phone:
                stats["skip"] += 1
                continue
            phone = str(phone).strip()
            ex = await session.execute(select(Lead).where(Lead.phone == phone))
            if ex.scalar_one_or_none():
                stats["dup"] += 1
                continue
            try:
                sp_n = row.get("איש מכירות ")
                sp_id = sp_ids.get(SP_MAP.get(sp_n.strip())) if sp_n else None
                c_n = row.get("מוצר שמתעניין")
                c_id = c_ids.get(C_MAP.get(c_n.strip())) if c_n else None
                resp = row.get("סטטוס מענה")
                session.add(Lead(
                    full_name=row.get("שם מלא") or "ליד ללא שם", phone=phone,
                    email=row.get("מייל לקוח"), city=row.get("עיר מגורים"),
                    address=row.get("כתובת"), notes=row.get("הערות ליד"),
                    source_type="ייבוא ממערכת ישנה", source_message=row.get("הודעה מהליד"),
                    source_name=row.get("שלוחת מוצר"), campaign_name=row.get("שם המפרסם"),
                    arrival_date=pd(row.get("תאריך יצירה")) or dt.now(),
                    last_contact_date=pd(row.get("תאריך פניה אחרונה")),
                    status=ST_MAP.get(row.get("סטאטוס ליד", "ליד חדש"), "ליד חדש"),
                    lead_response=R_MAP.get(resp.strip()) if resp else None,
                    salesperson_id=sp_id, course_id=c_id, created_by="import_script"
                ))
                stats["ok"] += 1
            except Exception as e:
                stats["err"] += 1
        await session.commit()
    return {"message": "ייבוא הושלם!", "results": stats, "total": sum(stats.values())}



# --- Unified Server: API + Frontend SPA ---
# Instead of a catch-all route (which conflicts with API routers),
# we use a 404 exception handler. This way:
#   - All API/webhook routes work normally via their routers
#   - Only REAL 404s on non-API paths get redirected to React SPA
if FRONTEND_DIR.exists():
    # Mount static assets (JS/CSS/Images generated by Vite)
    if (FRONTEND_DIR / "assets").exists():
        app.mount("/assets", StaticFiles(directory=FRONTEND_DIR / "assets"), name="assets")

    # Root route - serve React app
    @app.get("/", include_in_schema=False)
    async def serve_root():
        return FileResponse(FRONTEND_DIR / "index.html")

    # SPA fallback: catch 404s and serve index.html for frontend routes
    @app.exception_handler(404)
    async def spa_fallback(request: Request, exc: StarletteHTTPException):
        path = request.url.path
        # API and webhook 404s stay as real 404s
        if path.startswith("/api/") or path.startswith("/webhooks/"):
            return JSONResponse(status_code=404, content={"detail": str(exc.detail)})
        
        # For /assets/ paths, return real 404 if file doesn't exist
        # This prevents serving HTML for missing JS/CSS files
        if path.startswith("/assets/"):
            return JSONResponse(status_code=404, content={"detail": "Asset not found"})
        
        # Check if a physical file exists (favicon.ico, manifest.json, etc.)
        file_path = FRONTEND_DIR / path.lstrip("/")
        if file_path.exists() and file_path.is_file():
            return FileResponse(file_path)
        # Everything else: let React Router handle it
        return FileResponse(FRONTEND_DIR / "index.html")
else:
    @app.get("/")
    async def root():
        return {
            "status": "warning",
            "message": "Frontend not built. Run: cd frontend && npm run build",
            "api_status": "online"
        }

